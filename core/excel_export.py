"""
core/excel_export.py – Formatted Excel export for KIBAD.

Generates production-quality Excel workbooks with:
- Colored headers (bank-style blue/white)
- Freeze pane on row 1
- Auto column widths
- Number formatting (thousands, decimals)
- Totals row with bold formatting
- Conditional coloring for delta/change columns
- Named sheets
"""
from __future__ import annotations

import io
from typing import Any

import pandas as pd


# ---------------------------------------------------------------------------
# Colour palette (Sberbank-neutral: deep navy + white + accent green/red)
# ---------------------------------------------------------------------------
HEADER_FILL   = "1F3864"   # dark navy
HEADER_FONT   = "FFFFFF"   # white
ALT_ROW_FILL  = "EEF2F7"   # light blue-grey
TOTAL_FILL    = "D6E4F0"   # medium blue
TOTAL_FONT    = "1F3864"   # dark navy
POS_FILL      = "C6EFCE"   # light green
POS_FONT      = "276221"
NEG_FILL      = "FFC7CE"   # light red
NEG_FONT      = "9C0006"
NEUTRAL_FILL  = "FFEB9C"   # light yellow
NEUTRAL_FONT  = "9C6500"


def _openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def to_formatted_excel(
    df: pd.DataFrame,
    sheet_name: str = "Данные",
    title: str | None = None,
    totals_row: bool = True,
    delta_cols: list[str] | None = None,
    pct_cols: list[str] | None = None,
    int_cols: list[str] | None = None,
    freeze_cols: int = 0,
    extra_sheets: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    """Generate a formatted Excel workbook from a DataFrame.

    Parameters
    ----------
    df:
        Primary DataFrame to export.
    sheet_name:
        Name of the main worksheet.
    title:
        Optional title row to insert above the header.
    totals_row:
        If True, append a bolded ИТОГО row summing numeric columns.
    delta_cols:
        Column names that represent deltas — color positive green, negative red.
    pct_cols:
        Column names to format as percentages (0.XX → XX%).
    int_cols:
        Column names to format as integers (no decimals).
    freeze_cols:
        Number of columns to freeze from left (in addition to header row).
    extra_sheets:
        Additional DataFrames to include as separate sheets.

    Returns
    -------
    bytes
        Excel file content as bytes (ready for st.download_button).
    """
    if not _openpyxl_available():
        # Fallback: plain xlsx via pandas
        buf = io.BytesIO()
        df.to_excel(buf, index=False, sheet_name=sheet_name)
        return buf.getvalue()

    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    delta_cols = set(delta_cols or [])
    pct_cols = set(pct_cols or [])
    int_cols = set(int_cols or [])

    # --- Styles ---
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT, size=11)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_FILL)
    total_fill  = PatternFill("solid", fgColor=TOTAL_FILL)
    total_font  = Font(name="Calibri", bold=True, color=TOTAL_FONT, size=11)
    pos_fill    = PatternFill("solid", fgColor=POS_FILL)
    pos_font    = Font(name="Calibri", color=POS_FONT)
    neg_fill    = PatternFill("solid", fgColor=NEG_FILL)
    neg_font    = Font(name="Calibri", color=NEG_FONT)
    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    # Number formats
    FMT_NUMBER  = '#,##0.00'
    FMT_INT     = '#,##0'
    FMT_PCT     = '0.00%'
    FMT_TEXT    = '@'

    row_offset = 1

    # --- Optional title row ---
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=13, color=HEADER_FILL)
        ws.cell(row=1, column=1).alignment = left_align
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=max(len(df.columns), 1))
        row_offset = 2

    # --- Header row ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=row_offset, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = Border(bottom=Side(style="medium", color="FFFFFF"))

    # --- Data rows ---
    num_cols_in_df = df.select_dtypes(include="number").columns.tolist()
    col_names = list(df.columns)

    for r_idx, row in enumerate(df.itertuples(index=False), start=row_offset + 1):
        is_alt = (r_idx - row_offset) % 2 == 0
        for c_idx, col_name in enumerate(col_names, start=1):
            val = getattr(row, col_name.replace(" ", "_").replace("-", "_").replace("%", "_"), None)
            # Fallback: use positional access
            try:
                val = row[c_idx - 1]
            except Exception:
                pass

            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = right_align if col_name in num_cols_in_df else left_align

            # Alternating row colour (only if no delta colouring)
            if col_name not in delta_cols:
                if is_alt:
                    cell.fill = alt_fill

            # Number formatting
            if col_name in pct_cols:
                cell.number_format = FMT_PCT
            elif col_name in int_cols:
                cell.number_format = FMT_INT
            elif col_name in num_cols_in_df:
                cell.number_format = FMT_NUMBER

            # Delta colouring
            if col_name in delta_cols and val is not None:
                try:
                    fval = float(val)
                    if fval > 0:
                        cell.fill = pos_fill
                        cell.font = pos_font
                    elif fval < 0:
                        cell.fill = neg_fill
                        cell.font = neg_font
                except (TypeError, ValueError):
                    pass

    data_end_row = row_offset + len(df)

    # --- Totals row ---
    if totals_row and num_cols_in_df:
        total_row_idx = data_end_row + 1
        first_text_col = next((c for c in col_names if c not in num_cols_in_df), None)
        for c_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=total_row_idx, column=c_idx)
            if col_name == first_text_col:
                cell.value = "ИТОГО"
                cell.alignment = left_align
            elif col_name in num_cols_in_df and col_name not in delta_cols and col_name not in pct_cols:
                # SUM formula
                col_letter = get_column_letter(c_idx)
                cell.value = f"=SUM({col_letter}{row_offset + 1}:{col_letter}{data_end_row})"
                cell.number_format = FMT_INT if col_name in int_cols else FMT_NUMBER
                cell.alignment = right_align
            cell.fill = total_fill
            cell.font = total_font
            cell.border = Border(top=Side(style="medium", color=HEADER_FILL))

    # --- Column widths ---
    for c_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(c_idx)
        max_len = len(str(col_name))
        for val in df[col_name].head(200):
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # --- Freeze panes ---
    freeze_cell = ws.cell(row=row_offset + 1, column=freeze_cols + 1)
    ws.freeze_panes = freeze_cell

    # --- Auto filter ---
    if len(df) > 0:
        last_col = get_column_letter(len(col_names))
        ws.auto_filter.ref = f"A{row_offset}:{last_col}{row_offset}"

    # --- Extra sheets ---
    for sname, sdf in (extra_sheets or {}).items():
        ws2 = wb.create_sheet(title=sname[:31])
        # Simple export for extra sheets
        for c_idx, col_name in enumerate(sdf.columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=str(col_name))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        for r_idx, row in enumerate(dataframe_to_rows(sdf, index=False, header=False), start=2):
            for c_idx, val in enumerate(row, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def quick_excel(df: pd.DataFrame, sheet_name: str = "Данные") -> bytes:
    """Convenience wrapper: formatted Excel with sensible defaults."""
    # Auto-detect delta columns (containing 'изменен', 'delta', 'diff', '%')
    delta_cols = [
        c for c in df.columns
        if any(kw in c.lower() for kw in ("изменени", "delta", "diff", "отклон"))
    ]
    pct_cols = [
        c for c in df.columns
        if any(kw in c.lower() for kw in ("_%", "_pct", "_rate", "доля", "процент", "ratio"))
        and c not in delta_cols
    ]
    return to_formatted_excel(
        df,
        sheet_name=sheet_name,
        totals_row=True,
        delta_cols=delta_cols,
        pct_cols=pct_cols,
    )
